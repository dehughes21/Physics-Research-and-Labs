import numpy as np
import math
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.colors import BoundaryNorm
from matplotlib.ticker import MaxNLocator
import openpyxl
from openpyxl import load_workbook
import statistics
import datetime


def autoplotDataCain():
    fileName = input("Enter file name: ")
    time = []
    B_radial = []
    B_theta = []
    B_phi = []
    B_mag = []

    with open(fileName, 'r') as data:
        next(data)
        next(data)
        next(data)
        for line in data:
            line = line.split(',')
            time.append(line[0].strip())
            B_radial.append(float(line[1].strip()))
            B_phi.append(float(line[2].strip()))
            B_theta.append(float(line[3].strip()))
            B_mag.append(float(line[4].strip()))
    return time, B_radial, B_phi, B_theta, B_mag


def MagData_Morsch():
    magField = []
    magFieldR = []
    magFieldV = []
    magFieldLog = []
    magLocs = []

    magPath = 'MORSCH_B_FIELDS_CONTOUR_BR_BH.txt'

    with open(magPath, 'r') as mag:
        for line in mag:
            mag = float(line.split()[2])
            magR = float(line.split()[3])
            magV = magR / mag
            magLog = math.log(float(line.split()[2]), 10)

            lat = float(line.split()[0])
            long = float(line.split()[0])
            loc = (lat, long)

            magLocs.append(loc)
            magField.append(mag)
            magFieldR.append(magR)
            magFieldV.append(magV)
            magFieldLog.append(magLog)

    return magField, magFieldR, magFieldV, magFieldLog, magLocs


def columnData(path, shName, col):
    headerRows = int(input("How many rows is the header for column " + col + "? (input 0 if no header) "))
    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook[shName]
    dataList = []
    n = 0
    for cell in sheet[col]:
        if n > (headerRows - 1):
            dataList.insert(n, cell.value)
        n += 1
    dataList = [data for data in dataList if data is not None]
    return dataList


def columnData_sortASPERA(path, shName, col):
    headerRows = int(input("How many rows is the header for column " + col + "? (input 0 if no header) "))
    workbook = load_workbook(filename=path)
    sheet = workbook[shName]
    dataList_increase, dataList_decrease, dataList_nochange = [], [], []
    n = 0
    for cell in sheet[col]:
        if n > (headerRows - 1):
            Type = sheet.cell(row=n, column=26).value
            if Type == "I":
                dataList_increase.append(cell.value)
            elif Type == "N":
                dataList_nochange.append(cell.value)
            elif Type == "I":
                dataList_decrease.append(cell.value)
        n += 1
    dataList_increase = [data for data in dataList_increase if data is not None]
    dataList_nochange = [data for data in dataList_nochange if data is not None]
    dataList_decrease = [data for data in dataList_decrease if data is not None]

    return dataList_increase, dataList_nochange, dataList_decrease


def columnData_sortAltLim(path, shName, col):
    headerRows = int(input("How many rows is the header for column " + col + "? (input 0 if no header) "))
    altMin = float(input("Minimum altitude: "))
    altMax = float(input("Maximum altitude: "))
    workbook = load_workbook(filename=path)
    sheet = workbook[shName]
    dataList = []
    n = 0
    for cell in sheet[col]:
        if n > (headerRows - 1):
            if col in 'CDEFGH':
                alt = sheet.cell(row=n, column=4)
            elif col in "IJKLMN":
                alt = sheet.cell(row=n, column=10)
            elif col in "OPQRST":
                alt = sheet.cell(row=n, column=16)
            try:
                if (alt >= altMin) and (alt <= altMax):
                    dataList.append(cell.value)
            except UnboundLocalError:
                print("Pick a valid column")
                break
        n += 1
    dataList = [data for data in dataList if data is not None]
    return dataList


def columnData_sortSZA(path, shName, col):
    headerRows = int(input("How many rows is the header for column " + col + "? (input 0 if no header) "))
    SZAMin = float(input("Minimum SZA: "))
    SZAMax = float(input("Maximum SZA: "))
    workbook = load_workbook(filename=path)
    sheet = workbook[shName]
    dataList = []
    n = 0
    for cell in sheet[col]:
        if n > (headerRows - 1):
            if col in 'CDEFGH':
                SZA = sheet.cell(row=n, column=3)
            elif col in "IJKLMN":
                SZA = sheet.cell(row=n, column=9)
            elif col in "OPQRST":
                SZA = sheet.cell(row=n, column=15)

            if (SZA >= SZAMin) and (SZA <= SZAMax):
                dataList.append(cell.value)
        n += 1
    dataList = [data for data in dataList if data is not None]
    return dataList
